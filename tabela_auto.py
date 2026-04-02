import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, PatternFill
import pandas as pd

def estilizar(sheet, lista, estilo, tipo):
    """
    lista: ['A1', 'B1']
    estilo: o objeto de estilo (Font, PatternFill, etc.)
    tipo: string com o nome da propriedade ('font', 'fill', 'alignment', 'border')
    """

    for ref in lista:
        # setattr(objeto, nome_da_propriedade, valor)
        setattr(sheet[ref], tipo, estilo)

def header(ano, mes, moeda, smp):
    """
    Cria e formata o template inicial do Diário de Caixa em Excel.

    A função realiza a montagem estrutural (merge de células), define rótulos,
    aplica estilos de fonte Times New Roman, preenchimentos de cores (branco, 
    cinzento, laranja) e regras de bordas médias e finas.

    Dependências:
        Requer a função auxiliar estilizar(lista, objeto_estilo, tipo_estilo).
        
    Arquivo gerado:
        Custo_Entrada&Saida.xlsx
    """
    # __________Criar um novo arquivo excell e criar a planilha__________
    wb = openpyxl.Workbook()
    wb.create_sheet(f'Tabela Custo de Entrada&Saida-{mes}_{ano}')
    wb.remove(wb['Sheet'])


    # __________________Labels e Formatação________________
    # Labels
    sheet = wb[f'Tabela Custo de Entrada&Saida-{mes}_{ano}']
    sheet['F1'] = 'Ano:'
    sheet['F2'] = 'Mês:'
    sheet['F3'] = 'Moeda:'

    sheet['G1'] = ano
    sheet['G2'] = mes
    sheet['G3'] = moeda

    sheet['A5'] = 'DIÁRIO DE CAIXA'
    sheet.merge_cells('A5:G5')

    sheet['A6'] = 'Nº'
    sheet.merge_cells('A6:A7')
    sheet['B6'] = 'DATA'
    sheet.merge_cells('B6:B7')
    sheet['C6'] = 'DESIGNAÇÃO'
    sheet.merge_cells('C6:D7')
    sheet['E6'] = 'ENTRADAS (+)'
    sheet.merge_cells('E6:E7')
    sheet['F6'] = 'SAÍDAS (-)'
    sheet.merge_cells('F6:F7')
    sheet['G6'] = 'SALDO'
    sheet.merge_cells('G6:G7')

    sheet['D9'] = 'Saldo do Mês Anterior'
    sheet['G9'] = ' ' + str(smp)

    # Formatação
    celula = sheet['A5']
    celula.alignment = Alignment(horizontal='center', vertical='center')

    celula = sheet['A6']
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula = sheet['C6']
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula = sheet['B6']
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula = sheet['E6']
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula = sheet['F6']
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula = sheet['G6']
    celula.alignment = Alignment(horizontal='center', vertical='center')


    #____________________________Estilos________________________________
    # Estilos do textos
    sheet['F1'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet['F2'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet['F3'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")

    sheet['G1'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet['G2'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet['G3'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")


    sheet['A5'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")

    sheet['A6'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet['B6'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet['C6'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet['E6'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet['F6'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet['G6'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")

    sheet['D9'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet['E9'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")

    sheet['G9'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")

    # Estilos da planilha
    # Cores
    branco = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    cinzento = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    laranja = PatternFill(fill_type='solid', start_color='ffc000', end_color='ffc000')

    # Localizações de onde pintar
    brancos = ['A1', 'A2', 'A3', 'A4', 'A5', 'B1', 'B2', 'B3', 'B4', 'C1', 'C2', 'C3', 'C4',
                'D1', 'D2', 'D3', 'D4', 'E1', 'E2', 'E3', 'E4', 'F1', 'F2', 'F3', 'F4',
                'G1', 'G2', 'G3', 'G4']
    cinzentos = ['A6', 'B6', 'C6', 'D6', 'D9', 'E6', 'E9', 'F6', 'F9', 'G6', 'G9']
    laranjas = ['A8', 'A9', 'B8', 'B9', 'C8', 'C9', 'D8', 'E8', 'F8', 'G8']

    # Pintar nas localizações
    estilizar(sheet, brancos, branco, 'fill')
    estilizar(sheet, cinzentos, cinzento, 'fill')
    estilizar(sheet, laranjas, laranja, 'fill')

    # Bordas
    lista2 = ['D9', 'E9', 'F9', 'G9']
    lista3 = ['A6', 'B6', 'C6', 'E6', 'F6', 'G6', 'A7', 'B7', 'C7', 'E7', 'F7', 'G7']

    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    linha_media = openpyxl.styles.Side(color='000000', style='medium')
    linha_grossa = openpyxl.styles.Side(color='000000', style='thick')

    bordas_finas = openpyxl.styles.Border(left=linha_fina, right=linha_fina, top=linha_fina, bottom=linha_fina)
    bordas_medias = openpyxl.styles.Border(left=linha_media, right=linha_media, top=linha_media, bottom=linha_media)
    bordas_grossas = openpyxl.styles.Border(left=linha_grossa, right=linha_grossa, top=linha_grossa, bottom=linha_grossa)

    estilizar(sheet, ['G1', 'G2', 'G3'], bordas_finas, 'border')
    estilizar(sheet, lista3, bordas_medias, 'border')
    estilizar(sheet,lista2, openpyxl.styles.Border(top=linha_media), 'border')
    estilizar(sheet, ['D9'], openpyxl.styles.Border(left=linha_media, top=linha_media), 'border')
    estilizar(sheet, ['G9'], openpyxl.styles.Border(left=linha_media, right=linha_media, top=linha_media), 'border')

    # Este bloco deve estar dentro de uma função ou no final do arquivo antes do save
    for col in sheet.columns:
        max_length = 0  # <--- Note os 4 espaços de recuo aqui
        column = col[0].column_letter 
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # Salvar o arquivo
    wb.save('Custo_Entrada&Saida.xlsx')

def body(lista, mes, ano):
    """
    Função para adicionar itens ao Diário de Caixa, calculando o saldo de forma dinâmica.

    A função solicita ao usuário a entrada de dados para cada item (data, designação, entrada, saída),
    calcula o saldo atual com base no saldo anterior e adiciona o item à lista de itens. O processo
    continua até que o usuário decida parar.

    Dependências:
        Requer a função header() para criar o template inicial do Excel.
        
    Saída:
        Imprime a lista final de itens adicionados.
    """
 # Carregar o arquivo Excel existente e acessar a planilha correta
    wb = openpyxl.load_workbook('Custo_Entrada&Saida.xlsx')    
    sheet = wb[f'Tabela Custo de Entrada&Saida-{mes}_{ano}']
    cont_table = 10 # Contador para controlar a linha onde os itens serão inseridos, iniciando na linha 10 (após o saldo do mês anterior)

 # Iterar sobre a lista de itens e adicionar cada um à planilha, formatando as células conforme necessário
    for i in range(len(lista)):
        item = lista[i]
        #print(f'Nº: {item[0]}, Data: {item[1]}, Designação: {item[2]}, Entrada: {item[3]}, Saída: {item[4]}, Saldo: {item[5]}')
        sheet.merge_cells(f'C{cont_table}:D{cont_table}') # Mesclar as células C e D para a designação

    # Adicionar os valores do item às células correspondentes e aplicar formatação
        sheet[f'A{cont_table}'] = item[0]
        sheet[f'B{cont_table}'] = item[1]
        sheet[f'C{cont_table}'] = item[2]
        sheet.alignment = Alignment(horizontal='center', vertical='center')
        sheet[f'E{cont_table}'] = item[3]
        sheet[f'F{cont_table}'] = item[4]
        sheet[f'G{cont_table}'] = item[5]

    # Aplicar formatação de fonte e bordas às células do item
        sheet[f'A{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        sheet[f'B{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        sheet[f'C{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        sheet[f'E{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        sheet[f'F{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        sheet[f'G{cont_table}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
        
    # Aplicar bordas finas às células do item
        linha_fina = openpyxl.styles.Side(color='000000', style='thin')
        bordas_finas = openpyxl.styles.Border(left=linha_fina, right=linha_fina, top=linha_fina, bottom=linha_fina)
        sheet[f'A{cont_table}'].border = bordas_finas
        sheet[f'B{cont_table}'].border = bordas_finas
        sheet[f'C{cont_table}'].border = bordas_finas
        sheet[f'E{cont_table}'].border = bordas_finas
        sheet[f'F{cont_table}'].border = bordas_finas
        sheet[f'G{cont_table}'].border = bordas_finas
        cont_table += 1

    # Este bloco deve estar dentro de uma função ou no final do arquivo antes do save
    for col in sheet.columns:
        max_length = 0  # <--- Note os 4 espaços de recuo aqui
        column = col[0].column_letter 
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    # Salvar o arquivo
    wb.save('Custo_Entrada&Saida.xlsx')

def footer(lista, mes, ano, sexo, nome, saldo_atual, p_total, saldo_anterior, despesa_anterior, ativos_total):
    wb = openpyxl.load_workbook('Custo_Entrada&Saida.xlsx')    
    sheet = wb[f'Tabela Custo de Entrada&Saida-{mes}_{ano}']

    tamanho = len(lista) + 10 # Calcular o tamanho da tabela com base na quantidade de itens adicionados, considerando a linha inicial (10) e o saldo do mês anterior (linha 9)
    #print(f'{tamanho}')  Imprimir o tamanho da tabela para verificação

    branco = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
    cinzento = PatternFill(fill_type='solid', start_color='808080', end_color='808080')
    laranja = PatternFill(fill_type='solid', start_color='ffc000', end_color='ffc000')

    lista1 = [f'A{tamanho}', f'B{tamanho}', f'C{tamanho}', f'D{tamanho}', f'F{tamanho}', f'E{tamanho + 2}', f'F{tamanho + 2}', f'G{tamanho + 2}', f'G{tamanho + 4}']
    lista2 = [f'E{tamanho + 3}', f'F{tamanho + 3}', f'G{tamanho + 3}']

    estilizar(sheet, lista1, cinzento, 'fill')
    estilizar(sheet, [f'E{tamanho}', f'G{tamanho}'], branco, 'fill')
    estilizar(sheet, [f'A{tamanho + 1}', f'B{tamanho + 1}', f'C{tamanho + 1}', f'D{tamanho + 1}', f'E{tamanho + 1}', f'F{tamanho + 1}', f'G{tamanho + 1}'], branco, 'fill')
    estilizar(sheet, [f'D{tamanho + 2}', f'D{tamanho + 3}', f'D{tamanho + 4}', f'A{tamanho + 5}', f'A{tamanho + 6}', f'A{tamanho + 7}', f'B{tamanho + 5}', f'B{tamanho + 6}', f'B{tamanho + 7}', f'C{tamanho + 5}', f'C{tamanho + 6}', f'C{tamanho + 7}'], branco, 'fill')

    linha_fina = openpyxl.styles.Side(color='000000', style='thin')
    linha_media = openpyxl.styles.Side(color='000000', style='medium')
    bordas_medias = openpyxl.styles.Border(right=linha_media, left=linha_media, top=linha_media, bottom=linha_media)

    estilizar(sheet, lista2, laranja, 'fill')
    estilizar(sheet, lista1, bordas_medias, 'border')
    estilizar(sheet, lista2, bordas_medias, 'border')
    estilizar(sheet, [f'E{tamanho}', f'A{tamanho + 7}', f'B{tamanho + 7}', f'C{tamanho + 7}'], openpyxl.styles.Border(bottom=linha_fina), 'border')
    estilizar(sheet, [f'G{tamanho}'], openpyxl.styles.Border(right=linha_fina, bottom=linha_fina), 'border')

    sheet[f'D{tamanho + 2}'] = 'MOVIMENTOS DO MÊS'
    sheet[f'D{tamanho + 3}'] = 'SALDO DO MÊS ANTERIOR'
    sheet[f'D{tamanho + 4}'] = 'SALDO DO MÊS SEGUINTE'
    sheet[f'E{tamanho + 4}'] = '|------------»»»»»»»»»»»»'

    sheet[f'F{tamanho}'] = float(p_total)
    sheet[f'G{tamanho}'] = float(saldo_atual)
    sheet[f'E{tamanho + 2}'] = float(ativos_total)
    sheet[f'F{tamanho + 2}'] = float(p_total)
    sheet[f'G{tamanho + 2}'] = float(ativos_total - p_total)

    sheet[f'E{tamanho + 3}'] = float(saldo_anterior)
    sheet[f'F{tamanho + 3}'] = float(despesa_anterior)
    sheet[f'G{tamanho + 3}'] = float(saldo_anterior - despesa_anterior)

    sheet[f'G{tamanho + 4}'] = '+' + str(float(saldo_atual))

    sheet[f'E{tamanho + 2}'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet[f'F{tamanho + 2}'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet[f'G{tamanho + 2}'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet[f'E{tamanho + 3}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet[f'F{tamanho + 3}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet[f'G{tamanho + 3}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet[f'G{tamanho + 4}'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet[f'F{tamanho}'].font = Font(name='Times New Roman', size=11, bold=True, color="ffffff")
    sheet[f'G{tamanho}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")

    sheet.merge_cells(f'A{tamanho}:D{tamanho}')
    sheet.merge_cells(f'A{tamanho + 5}:B{tamanho + 5}')
    sheet.merge_cells(f'A{tamanho + 6}:B{tamanho + 6}')

    if sexo.upper() == 'F':
        sheet[f'A{tamanho + 5}'] = 'A TESOUREIRA'
    elif sexo.upper() == 'M':
        sheet[f'A{tamanho + 5}'] = 'O TESOUREIRO'
    else:
        sheet[f'A{tamanho + 5}'] = 'O/A TESOUREIRO/A'

    sheet[f'A{tamanho + 6}'] = str(nome)

    sheet[f'D{tamanho + 2}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet[f'D{tamanho + 3}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet[f'D{tamanho + 4}'].font = Font(name='Times New Roman', size=11, bold=True, color="000000")
    sheet[f'A{tamanho + 5}'].font = Font(name='Times New Roman', size=11, bold=False, color="000000")
    sheet[f'A{tamanho + 6}'].font = Font(name='Times New Roman', size=11, bold=False, color="000000")

    sheet.column_dimensions['A'].width = 5   # Nº
    sheet.column_dimensions['B'].width = 15  # Data
    sheet.column_dimensions['C'].width = 15  # Designação (Parte 1)
    sheet.column_dimensions['D'].width = 35  # Designação (Parte 2)
    sheet.column_dimensions['E'].width = 20  # Entradas
    sheet.column_dimensions['F'].width = 18  # Saídas
    sheet.column_dimensions['G'].width = 20  # Saldo

    # Salvar o arquivo
    wb.save('Custo_Entrada&Saida.xlsx')
        
# Solicitar ao usuário o saldo do mês anterior, ano, mês e moeda para configurar o template do Diário de Caixa
#smp = float(input('Saldo do Mês Antrior, prime ENTER caso seja 0 > '))
#des_ant = float(input('Depesa total do mês passado> '))
#ano = input('Ano, ex: 2026 > ')
#mes = input('Mês, ex: JANEIRO > ').upper()
#saldo_atual = 0
#p_total = 0.0
#ativos_total = 0.0
#header(ano, mes, input('Moeda, ex: AKZ > ').upper(), float(smp - des_ant)) # Criar o template do Diário de Caixa com as informações fornecidas

#lista_itens = []  # Lista para armazenar os itens adicionados, cada item é uma lista: [Nº, Data, Designação, Entrada, Saída, Saldo]

# Loop para adicionar itens ao Diário de Caixa, solicitando ao usuário os detalhes de cada item e calculando o saldo atual
#while True:
# Solicitar ao usuário se deseja adicionar um novo item à tabela. Se sim, solicitar os detalhes do item e calcular o saldo atual com base no saldo do mês anterior e no saldo do último item adicionado.
#    if input('Adicionar um novo item a tabela? (S/n)> ').upper() == 'S':
#        data = input('Data (dd-mm-aaaa)> ')
#        designacao = input('Designação> ')
#        entrada = float(input('Valor de Entrada (0 se não houver)> '))
#        saida = float(input('Valor de Saída (0 se não houver)>  '))

    # Calcular o saldo atual com base no saldo do mês anterior, saldo do último item e as entradas/saídas do item atual
 #       if len(lista_itens) == 0:
  #          saldo_atual = smp - des_ant + entrada - saida

   #     else:
    #        saldo_anterior = lista_itens[-1][5]  # Saldo do último item
     #       saldo_atual = saldo_anterior + entrada - saida

      #  novo_item = [len(lista_itens) + 1, data, designacao, entrada, saida, saldo_atual]
       # lista_itens.append(novo_item)
       # print(f'Item adicionado: {novo_item}')

        #p_total = float(0)  # Variável para acumular o total de entradas e saídas
       # ativos_total = float(0)
       # for item in lista_itens:
        #    p_total += float(item[4])
#         3   ativos_total += float(item[3])

   # else:
        #print('Encerrando a adição de itens. Lista final de itens:\n')
    #    body(lista_itens, mes, ano) # Adicionar os itens à planilha do Diário de Caixa e salvar o arquivo Excel
     #   footer(lista_itens, mes, ano, input('Sexo (M/F)> '), input('Nome> '), saldo_atual, p_total, smp, des_ant, ativos_total) # Adicionar o rodapé à planilha do 	 
      #  input("\nProcesso concluído! Pressione ENTER para sair...") # Adicione esta linha
        #Diário de Caixa e salvar o arquivo Excel
        #for item in lista_itens:
         #   print(item)
       # break
